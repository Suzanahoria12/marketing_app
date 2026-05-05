import os
import random
import tkinter as tk
from tkinter import ttk, messagebox
import sys

try:
    import openpyxl
except ImportError:
    messagebox.showerror('Error', 'Falta la librería \'openpyxl\'.\nInstálala con: pip install openpyxl')
    sys.exit(1)


class QuizApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.geometry('850x850')  # Aumentado un poco el alto para las imágenes
        self.configure(bg='#F5F5F7')
        self.title("Quiz de Relaciones Públicas")
        self.style = ttk.Style(self)
        self.style.theme_use('clam')
        self.style.configure('TButton', font=('Helvetica', 11), padding=10)
        self.style.configure('TLabel', font=('Helvetica', 12), background='#F5F5F7')
        self.style.configure('Title.TLabel', font=('Helvetica', 20, 'bold'), background='#F5F5F7')

        self.preguntas = []
        self.indice_actual = 0
        self.puntuacion = 0
        self.mapa_relacion = {}
        self.sel_izq = None
        self.colores_pares = ['#FFD700', '#FFB6C1', '#98FB98', '#87CEFA', '#DDA0DD', '#F0E68C', '#E0FFFF', '#FFAB91']
        self.color_idx = 0

        self.main_frame = tk.Frame(self, bg='#F5F5F7')
        self.main_frame.pack(expand=True, fill=tk.BOTH, padx=40, pady=40)
        self.mostrar_menu_asignaturas()

    def limpiar_pantalla(self):
        for widget in self.main_frame.winfo_children():
            widget.destroy()

    def mostrar_menu_asignaturas(self):
        self.limpiar_pantalla()
        ttk.Label(self.main_frame, text='Elige una Asignatura', style='Title.TLabel').pack(pady=(0, 20))
        archivos_xlsx = [f for f in os.listdir('.') if f.endswith('.xlsx') and not f.startswith('~')]
        if not archivos_xlsx:
            ttk.Label(self.main_frame, text='No hay archivos .xlsx', foreground='red').pack()
            return
        for archivo in archivos_xlsx:
            ttk.Button(self.main_frame, text=os.path.splitext(archivo)[0],
                       command=lambda a=archivo: self.mostrar_menu_temas(a)).pack(fill=tk.X, pady=5)

    def mostrar_menu_temas(self, archivo):
        self.limpiar_pantalla()
        try:
            wb = openpyxl.load_workbook(archivo, data_only=True, read_only=True)
            nombres_hojas = wb.sheetnames
            wb.close()
        except:
            self.mostrar_menu_asignaturas();
            return
        ttk.Label(self.main_frame, text=f'Temas de {os.path.splitext(archivo)[0]}', style='Title.TLabel').pack(
            pady=(0, 20))
        for hoja in nombres_hojas:
            ttk.Button(self.main_frame, text=hoja,
                       command=lambda a=archivo, h=hoja: self.cargar_preguntas(a, h)).pack(fill=tk.X, pady=5)
        ttk.Button(self.main_frame, text='← Volver', command=self.mostrar_menu_asignaturas).pack(pady=20)

    def cargar_preguntas(self, archivo, hoja):
        try:
            wb = openpyxl.load_workbook(archivo, data_only=True)
            sheet = wb[hoja]
            headers = [str(cell.value).strip() if cell.value else '' for cell in sheet[1]]
            self.preguntas = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                fila = {headers[i]: str(row[i]).strip() if row[i] is not None else '' for i in range(len(headers)) if
                        headers[i]}
                if fila.get('PREGUNTA'): self.preguntas.append(fila)
            wb.close()
            random.shuffle(self.preguntas)
            self.indice_actual = 0;
            self.puntuacion = 0
            self.mostrar_pregunta()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def mostrar_pregunta(self):
        self.limpiar_pantalla()
        p = self.preguntas[self.indice_actual]
        tipo = p.get('TIPO', 'individual').lower()
        self.color_idx = 0

        ttk.Label(self.main_frame, text=f'Pregunta {self.indice_actual + 1} de {len(self.preguntas)}',
                  foreground='#888').pack(anchor=tk.W)
        tk.Label(self.main_frame, text=p['PREGUNTA'], font=('Helvetica', 14, 'bold'), bg='#F5F5F7', wraplength=750,
                 justify=tk.LEFT).pack(anchor=tk.W, pady=(10, 20))

        self.opciones_container = tk.Frame(self.main_frame, bg='#F5F5F7')
        self.opciones_container.pack(fill=tk.BOTH, expand=True)

        if tipo == 'relacionar':
            self.setup_relacionar(p)
        else:
            self.setup_seleccion(p, tipo == 'multiple')

        self.btn_validar = ttk.Button(self.main_frame, text='Validar Respuesta', command=lambda: self.validar(p))
        self.btn_validar.pack(pady=20)
        self.lbl_fb = tk.Label(self.main_frame, text='', font=('Helvetica', 12, 'bold'), bg='#F5F5F7', wraplength=750)
        self.lbl_fb.pack()
        self.lbl_just = tk.Label(self.main_frame, text='', font=('Helvetica', 11, 'italic'), bg='#F5F5F7', fg='#555',
                                 wraplength=750, justify=tk.LEFT)
        self.lbl_just.pack(pady=10)
        self.btn_sig = ttk.Button(self.main_frame, text='Siguiente →', state=tk.DISABLED, command=self.siguiente)
        self.btn_sig.pack(anchor=tk.E)

    def setup_seleccion(self, p, es_multiple):
        opciones = [p[f'OPCION {l}'] for l in ['A', 'B', 'C', 'D', 'E', 'F', 'G'] if p.get(f'OPCION {l}')]
        random.shuffle(opciones)
        self.widgets_opciones = []
        if es_multiple:
            self.vars_mul = []
            for opt in opciones:
                v = tk.StringVar(value="")
                c = ttk.Checkbutton(self.opciones_container, text=opt, variable=v, onvalue=opt, offvalue="")
                c.pack(anchor=tk.W, pady=2)
                self.widgets_opciones.append(c);
                self.vars_mul.append(v)
        else:
            self.radio_var = tk.StringVar(value="")
            for opt in opciones:
                r = ttk.Radiobutton(self.opciones_container, text=opt, value=opt, variable=self.radio_var)
                r.pack(anchor=tk.W, pady=2)
                self.widgets_opciones.append(r)

    def setup_relacionar(self, p):
        self.mapa_relacion = {};
        self.sel_izq = None
        pares = [par.split(' - ') for par in p['CORRECTA'].split(';') if ' - ' in par]
        izq_list = [par[0].strip() for par in pares]
        der_list = list(set(par[1].strip() for par in pares))
        random.shuffle(izq_list);
        random.shuffle(der_list)

        self.f_izq = tk.Frame(self.opciones_container, bg='#F5F5F7');
        self.f_izq.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.f_der = tk.Frame(self.opciones_container, bg='#F5F5F7');
        self.f_der.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        self.btns_izq = {};
        self.btns_der = {}
        for txt in izq_list:
            btn = tk.Button(self.f_izq, text=txt, wraplength=300, relief=tk.GROOVE, bg='white',
                            command=lambda t=txt: self.click_izq(t))
            btn.pack(fill=tk.X, padx=5, pady=2);
            self.btns_izq[txt] = btn
        for txt in der_list:
            btn = tk.Button(self.f_der, text=txt, wraplength=300, relief=tk.GROOVE, bg='white',
                            command=lambda t=txt: self.click_der(t))
            btn.pack(fill=tk.X, padx=5, pady=2);
            self.btns_der[txt] = btn

    def click_izq(self, txt):
        if self.sel_izq: self.btns_izq[self.sel_izq].config(bg='white')
        self.sel_izq = txt
        self.btns_izq[txt].config(bg='#ADD8E6')

    def click_der(self, txt):
        if not self.sel_izq: return
        color = self.colores_pares[self.color_idx % len(self.colores_pares)]
        self.mapa_relacion[self.sel_izq] = txt
        self.btns_izq[self.sel_izq].config(bg=color)
        self.btns_der[txt].config(bg=color)
        self.color_idx += 1
        self.sel_izq = None

    def validar(self, p):
        tipo = p.get('TIPO', 'individual').lower()
        if tipo == 'relacionar':
            res_usuario = set(f"{k} - {v}" for k, v in self.mapa_relacion.items())
            res_correcta = set(par.strip() for par in p['CORRECTA'].split(';'))
            es_ok = res_usuario == res_correcta
        elif tipo == 'multiple':
            sel = [v.get() for v in self.vars_mul if v.get()]
            es_ok = set(sel) == set(par.strip() for par in p['CORRECTA'].split(';'))
        else:
            es_ok = self.radio_var.get() == p['CORRECTA']

        self.btn_validar.config(state=tk.DISABLED);
        self.btn_sig.config(state=tk.NORMAL)
        if es_ok:
            self.lbl_fb.config(text="✓ CORRECTO", fg="#28a745");
            self.puntuacion += 1
        else:
            self.lbl_fb.config(text=f"✗ INCORRECTO. Era:\n{p['CORRECTA']}", fg="#dc3545")
        if p.get('JUSTIFICACION'): self.lbl_just.config(text=f"Justificación: {p['JUSTIFICACION']}")

    def siguiente(self):
        self.indice_actual += 1
        if self.indice_actual < len(self.preguntas):
            self.mostrar_pregunta()
        else:
            self.mostrar_resultados()

    def mostrar_resultados(self):
        self.limpiar_pantalla()
        total = len(self.preguntas)
        nota = (self.puntuacion / total) * 100 if total > 0 else 0

        # --- LÓGICA DE IMAGEN ---
        img_name = "1.png" if nota >= 70 else "2.png"

        tk.Label(self.main_frame, text="Examen Finalizado", font=('Helvetica', 18, 'bold'), bg='#F5F5F7').pack(pady=10)

        # Mostrar imagen
        try:
            self.img_final = tk.PhotoImage(file=img_name)
            lbl_img = tk.Label(self.main_frame, image=self.img_final, bg='#F5F5F7')
            lbl_img.pack(pady=10)
        except Exception as e:
            tk.Label(self.main_frame, text=f"(Imagen {img_name} no encontrada)", fg="red", bg='#F5F5F7').pack()

        tk.Label(self.main_frame, text=f"Nota: {nota:.1f}% ({self.puntuacion}/{total})", font=('Helvetica', 16),
                 bg='#F5F5F7').pack(pady=10)

        ttk.Button(self.main_frame, text="Inicio", command=self.mostrar_menu_asignaturas).pack(pady=20)


if __name__ == '__main__':
    QuizApp().mainloop()
